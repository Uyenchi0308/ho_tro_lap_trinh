import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import openpyxl
import pytz
from telegram import Update
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    ConversationHandler,
    filters
)
import logging

# Cấu hình logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Cấu hình timezone
try:
    import tzlocal

    TIMEZONE = tzlocal.get_localzone()
except:
    TIMEZONE = pytz.timezone('Asia/Ho_Chi_Minh')

# Quy định màu sắc cho từng sợi cáp quang (1-24)
FIBER_COLORS = {
    1: ('Xanh dương', '0000FF'),
    2: ('Cam', 'FFA500'),
    3: ('Xanh lá', '00FF00'),
    4: ('Nâu', 'A52A2A'),
    5: ('Xám', '808080'),
    6: ('Trắng', 'FFFFFF'),
    7: ('Đỏ', 'FF0000'),
    8: ('Đen', '000000'),
    9: ('Vàng', 'FFFF00'),
    10: ('Tím', '800080'),
    11: ('Hồng', 'FFC0CB'),
    12: ('Xanh ngọc', '00FFFF'),
    13: ('Xanh dương', '0000FF'),
    14: ('Cam', 'FFA500'),
    15: ('Xanh lá', '00FF00'),
    16: ('Nâu', 'A52A2A'),
    17: ('Xám', '808080'),
    18: ('Trắng', 'FFFFFF'),
    19: ('Đỏ', 'FF0000'),
    20: ('Đen', '000000'),
    21: ('Vàng', 'FFFF00'),
    22: ('Tím', '800080'),
    23: ('Hồng', 'FFC0CB'),
    24: ('Xanh ngọc', '00FFFF')
}

# Dictionary mẫu cho co nhiệt
HEAT_SHRINKS = {
    'HS-1': [1, 2],
    'HS-2': [3, 4],
    'HS-3': [5, 6],
    'HS-4': [7, 8],
    'HS-5': [9, 10],
    'HS-6': [11, 12],
    'HS-7': [13, 14],
    'HS-8': [15, 16],
    'HS-9': [17, 18],
    'HS-10': [19, 20],
    'HS-11': [21, 22],
    'HS-12': [23, 24]
}

# Kết nối mẫu
CONNECTIONS = {
    'MX1': {
        'location': {'lat': 10.12345, 'long': 106.12345},
        'connections': {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8,
                        9: 9, 10: 10, 11: 11, 12: 12, 13: 13, 14: 14,
                        15: 15, 16: 16, 17: 17, 18: 18, 19: 19, 20: 20,
                        21: 21, 22: 22, 23: 23, 24: 24}
    },
    'MX2': {
        'location': {'lat': 10.22345, 'long': 106.22345},
        'connections': {1: 2, 2: 1, 3: 4, 4: 3, 5: 6, 6: 5, 7: 8, 8: 7,
                        9: 10, 10: 9, 11: 12, 12: 11, 13: 14, 14: 13,
                        15: 16, 16: 15, 17: 18, 18: 17, 19: 20, 20: 19,
                        21: 22, 22: 21, 23: 24, 24: 23}
    }
}

# File Excel phân quyền
PERMISSION_FILE = 'quyen.xlsx'
# File Excel chính
MAIN_EXCEL_FILE = 'mang_xong_cap_quang.xlsx'

# Thêm trạng thái mới vào các biến trạng thái hiện có
FIND_MX, GET_MX, ADD_MX_NAME, ADD_MX_CONNECTIONS, EDIT_MX, EDIT_MX_CONNECTION = range(6)


def create_excel_file(filename=None):
    """Tạo file Excel mẫu cho quản lý măng xông cáp quang (phiên bản đồng bộ)"""
    try:
        # Sử dụng filename mặc định nếu không được cung cấp
        if filename is None:
            filename = MAIN_EXCEL_FILE

        # Lấy đường dẫn tuyệt đối
        abs_path = os.path.abspath(filename)
        dir_path = os.path.dirname(abs_path)

        # Đảm bảo thư mục tồn tại
        os.makedirs(dir_path, exist_ok=True)

        logger.info(f"Đang tạo file Excel tại: {abs_path}")
        print(f"Đang tạo file Excel tại: {abs_path}")

        wb = Workbook()

        # Tạo sheet cho từng măng xông
        for mx_name, mx_data in CONNECTIONS.items():
            ws = wb.create_sheet(title=mx_name)

            # Thêm thông tin vị trí
            ws['A1'] = 'Tên măng xông:'
            ws['B1'] = mx_name
            ws['A2'] = 'Vị trí (lat):'
            ws['B2'] = mx_data['location']['lat']
            ws['A3'] = 'Vị trí (long):'
            ws['B3'] = mx_data['location']['long']

            # Tiêu đề các cột
            headers = ['STT', 'Màu sắc', 'Co nhiệt', 'Vị trí trong co', 'Đầu vào', 'Đầu ra', 'Ghi chú']
            ws.append(headers)

            # Định dạng tiêu đề
            for col in range(1, len(headers) + 1):
                ws.cell(row=4, column=col).font = Font(bold=True)

            # Thêm dữ liệu cho từng sợi
            for fiber_num in range(1, 25):
                color_name, color_hex = FIBER_COLORS[fiber_num]

                # Tìm co nhiệt chứa sợi này
                hs_name = ''
                hs_pos = ''
                for hs, fibers in HEAT_SHRINKS.items():
                    if fiber_num in fibers:
                        hs_name = hs
                        pos = fibers.index(fiber_num) + 1
                        hs_pos = f"{pos}/{len(fibers)}"
                        break

                # Xác định đầu ra
                output_fiber = mx_data['connections'].get(fiber_num, fiber_num)

                # Xác định ghi chú
                note = 'Đấu thẳng' if fiber_num == output_fiber else 'Đấu chéo'

                # Thêm dòng dữ liệu
                ws.append([
                    fiber_num,
                    color_name,
                    hs_name,
                    hs_pos,
                    fiber_num,
                    output_fiber,
                    note
                ])

            # Định dạng màu cho các ô
            for row in range(5, 29):  # Dòng 5-28 tương ứng với sợi 1-24
                # Màu sợi cáp
                fiber_num = ws.cell(row=row, column=1).value
                if fiber_num:
                    _, color_hex = FIBER_COLORS[fiber_num]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    ws.cell(row=row, column=2).fill = fill

                    # Màu chữ (đen hoặc trắng tùy vào màu nền)
                    text_color = '000000' if color_hex in ['FFFFFF', '00FFFF', 'FFFF00'] else 'FFFFFF'
                    ws.cell(row=row, column=2).font = Font(color=text_color)

                # Định dạng có điều kiện cho cột đầu vào và đầu ra
                for col in [5, 6]:  # Cột E (5) và F (6)
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        _, color_hex = FIBER_COLORS[cell.value]
                        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                        cell.fill = fill
                        text_color = '000000' if color_hex in ['FFFFFF', '00FFFF', 'FFFF00'] else 'FFFFFF'
                        cell.font = Font(color=text_color)

            # Thiết lập Data Validation cho cột đầu vào và đầu ra
            dv = openpyxl.worksheet.datavalidation.DataValidation(
                type="whole",
                operator="between",
                formula1="1",
                formula2="24",
                showErrorMessage=True,
                errorTitle="Giá trị không hợp lệ",
                error="Vui lòng nhập số từ 1 đến 24"
            )
            ws.add_data_validation(dv)
            dv.add('E5:E28')  # Cột Đầu vào
            dv.add('F5:F28')  # Cột Đầu ra

            # Đặt chiều rộng cột
            column_widths = {'A': 8, 'B': 12, 'C': 10, 'D': 12, 'E': 10, 'F': 10, 'G': 15}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

        # Xóa sheet mặc định
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Lưu file với tên chính xác
        wb.save(filename)
        logger.info(f"Đã tạo file Excel thành công tại: {abs_path}")
        print(f"Đã tạo file Excel thành công tại: {abs_path}")
        return abs_path
    except PermissionError as e:
        error_msg = f"Lỗi quyền khi lưu file tại {abs_path}: {e}"
        logger.error(error_msg)
        print(error_msg)
        raise Exception("Không có quyền ghi file. Vui lòng kiểm tra quyền thư mục.")
    except Exception as e:
        error_msg = f"Lỗi khi tạo file Excel tại {abs_path}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        print(error_msg)
        raise Exception(f"Có lỗi xảy ra khi tạo file Excel: {str(e)}")


def check_permission(username, permission_type='write'):
    """Kiểm tra quyền của user"""
    if not os.path.exists(PERMISSION_FILE):
        return False

    try:
        df = pd.read_excel(PERMISSION_FILE)
        if permission_type == 'write':
            return username in df['username'].values
        return True
    except Exception as e:
        logger.error(f"Error reading permission file: {e}")
        return False


def find_mx_location(mx_name):
    """Tìm vị trí của măng xông"""
    return CONNECTIONS.get(mx_name.upper(), {}).get('location', None)


def get_mx_connections(mx_name):
    """Lấy thông tin đấu nối của măng xông"""
    return CONNECTIONS.get(mx_name.upper(), {}).get('connections', None)


def add_new_mx(mx_name, lat, long, connections):
    """Thêm măng xông mới vào hệ thống"""
    try:
        mx_name = mx_name.upper()
        if mx_name in CONNECTIONS:
            return False

        CONNECTIONS[mx_name] = {
            'location': {'lat': lat, 'long': long},
            'connections': connections
        }

        # Cập nhật file Excel
        update_excel_with_new_mx(mx_name, lat, long, connections)
        return True
    except Exception as e:
        logger.error(f"Error in add_new_mx: {e}")
        return False


def update_excel_with_new_mx(mx_name, lat, long, connections):
    """Cập nhật file Excel với măng xông mới"""
    try:
        # Mở file Excel hiện có
        if not os.path.exists(MAIN_EXCEL_FILE):
            create_excel_file(MAIN_EXCEL_FILE)

        wb = openpyxl.load_workbook(MAIN_EXCEL_FILE)

        # Tạo sheet mới cho măng xông
        ws = wb.create_sheet(title=mx_name)

        # Thêm thông tin vị trí
        ws['A1'] = 'Tên măng xông:'
        ws['B1'] = mx_name
        ws['A2'] = 'Vị trí (lat):'
        ws['B2'] = lat
        ws['A3'] = 'Vị trí (long):'
        ws['B3'] = long

        # Tiêu đề các cột
        headers = ['STT', 'Màu sắc', 'Co nhiệt', 'Vị trí trong co', 'Đầu vào', 'Đầu ra', 'Ghi chú']
        ws.append(headers)

        # Định dạng tiêu đề
        for col in range(1, len(headers) + 1):
            ws.cell(row=4, column=col).font = Font(bold=True)

        # Thêm dữ liệu cho từng sợi
        for fiber_num in range(1, 25):
            color_name, color_hex = FIBER_COLORS[fiber_num]

            # Tìm co nhiệt chứa sợi này
            hs_name = ''
            hs_pos = ''
            for hs, fibers in HEAT_SHRINKS.items():
                if fiber_num in fibers:
                    hs_name = hs
                    pos = fibers.index(fiber_num) + 1
                    hs_pos = f"{pos}/{len(fibers)}"
                    break

            # Xác định đầu ra
            output_fiber = connections.get(fiber_num, fiber_num)

            # Xác định ghi chú
            note = 'Đấu thẳng' if fiber_num == output_fiber else 'Đấu chéo'

            # Thêm dòng dữ liệu
            ws.append([
                fiber_num,
                color_name,
                hs_name,
                hs_pos,
                fiber_num,
                output_fiber,
                note
            ])

            # Định dạng màu cho các ô
            row = fiber_num + 4  # Dòng bắt đầu từ 5
            # Màu sợi cáp
            _, color_hex = FIBER_COLORS[fiber_num]
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            ws.cell(row=row, column=2).fill = fill

            # Màu chữ (đen hoặc trắng tùy vào màu nền)
            text_color = '000000' if color_hex in ['FFFFFF', '00FFFF', 'FFFF00'] else 'FFFFFF'
            ws.cell(row=row, column=2).font = Font(color=text_color)

            # Định dạng có điều kiện cho cột đầu vào và đầu ra
            for col_num in [5, 6]:  # Cột E (5) và F (6)
                cell = ws.cell(row=row, column=col_num)
                if cell.value:
                    _, color_hex = FIBER_COLORS[cell.value]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
                    cell.fill = fill
                    text_color = '000000' if color_hex in ['FFFFFF', '00FFFF', 'FFFF00'] else 'FFFFFF'
                    cell.font = Font(color=text_color)

        # Thiết lập Data Validation cho cột đầu vào và đầu ra
        dv = openpyxl.worksheet.datavalidation.DataValidation(
            type="whole",
            operator="between",
            formula1="1",
            formula2="24",
            showErrorMessage=True,
            errorTitle="Giá trị không hợp lệ",
            error="Vui lòng nhập số từ 1 đến 24"
        )
        ws.add_data_validation(dv)
        dv.add('E5:E28')  # Cột Đầu vào
        dv.add('F5:F28')  # Cột Đầu ra

        # Đặt chiều rộng cột
        column_widths = {'A': 8, 'B': 12, 'C': 10, 'D': 12, 'E': 10, 'F': 10, 'G': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Lưu file
        wb.save(MAIN_EXCEL_FILE)
        logger.info(f"Đã cập nhật file Excel với măng xông mới {mx_name}")

    except Exception as e:
        logger.error(f"Error updating Excel with new MX: {e}")
        raise

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /start"""
    try:
        user = update.effective_user
        await update.message.reply_text(
            f"Xin chào {user.first_name}!\n\n"
            "Đây là bot quản lý măng xông cáp quang. Các lệnh có sẵn:\n"
            "/start - Hiển thị thông tin này\n"
            "/help - Hướng dẫn sử dụng\n"
            "/findmx - Tìm vị trí măng xông\n"
            "/getmx - Xem thông tin đấu nối măng xông\n"
            "/addmx - Thêm măng xông mới (cần quyền ghi)\n"
            "/editmx - Sửa đấu nối măng xông (cần quyền ghi)\n"  # Thêm dòng mới
            "/download - Tải file Excel tổng hợp"
        )
    except Exception as e:
        logger.error(f"Error in start command: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý lệnh /start.")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh /help"""
    try:
        await update.message.reply_text(
            "HƯỚNG DẪN SỬ DỤNG:\n\n"
            "1. Tìm vị trí măng xông:\n"
            "   Gõ /findmx sau đó nhập tên măng xông (ví dụ: MX1)\n\n"
            "2. Xem thông tin đấu nối:\n"
            "   Gõ /getmx sau đó nhập tên măng xông\n\n"
            "3. Thêm măng xông mới (cần quyền):\n"
            "   Gõ /addmx và làm theo hướng dẫn\n\n"
            "4. Sửa đấu nối măng xông (cần quyền):\n"  # Thêm mục mới
            "   Gõ /editmx sau đó nhập tên măng xông và các cặp đấu nối cần sửa\n\n"
            "5. Tải file Excel tổng hợp:\n"
            "   Gõ /download để nhận file mới nhất"
        )
    except Exception as e:
        logger.error(f"Error in help command: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý lệnh /help.")


async def find_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh tìm măng xông"""
    try:
        await update.message.reply_text(
            "Vui lòng nhập tên măng xông cần tìm (ví dụ: MX1):"
        )
        return FIND_MX
    except Exception as e:
        logger.error(f"Error in find_mx command: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý lệnh /findmx.")
        return ConversationHandler.END


async def handle_find_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý tên măng xông được nhập"""
    try:
        mx_name = update.message.text.upper()
        location = find_mx_location(mx_name)

        if location:
            await update.message.reply_text(
                f"Vị trí măng xông {mx_name}:\n"
                f"Latitude: {location['lat']}\n"
                f"Longitude: {location['long']}\n\n"
                "Bạn có thể copy toạ độ này để sử dụng."
            )
        else:
            await update.message.reply_text(
                f"Không tìm thấy măng xông {mx_name} trong hệ thống."
            )
    except Exception as e:
        logger.error(f"Error in handle_find_mx: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý yêu cầu tìm măng xông.")

    return ConversationHandler.END


async def get_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh xem thông tin măng xông"""
    try:
        await update.message.reply_text(
            "Vui lòng nhập tên măng xông cần xem thông tin đấu nối (ví dụ: MX1):"
        )
        return GET_MX
    except Exception as e:
        logger.error(f"Error in get_mx command: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý lệnh /getmx.")
        return ConversationHandler.END


async def handle_get_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý thông tin đấu nối măng xông"""
    try:
        mx_name = update.message.text.upper()
        connections = get_mx_connections(mx_name)

        if not connections:
            await update.message.reply_text(f"Không tìm thấy măng xông {mx_name} trong hệ thống.")
            return ConversationHandler.END

        # Tạo thông điệp hiển thị
        message = f"Thông tin đấu nối măng xông {mx_name}:\n\n"
        message += "Sợi | Đầu vào -> Đầu ra | Ghi chú\n"
        message += "---------------------------\n"

        for input_fiber, output_fiber in connections.items():
            note = "Thẳng" if input_fiber == output_fiber else "Chéo"
            color_name, _ = FIBER_COLORS[input_fiber]
            message += f"{input_fiber:2} ({color_name:10}) -> {output_fiber:2} | {note}\n"

        await update.message.reply_text(message)
    except Exception as e:
        logger.error(f"Error in handle_get_mx: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý yêu cầu xem măng xông.")

    return ConversationHandler.END


async def add_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh thêm măng xông mới"""
    try:
        user = update.effective_user

        # Kiểm tra quyền
        if not check_permission(user.username, 'write'):
            await update.message.reply_text(
                "Bạn không có quyền thêm măng xông mới. "
                "Liên hệ quản trị viên để được cấp quyền."
            )
            return ConversationHandler.END

        await update.message.reply_text(
            "Vui lòng nhập thông tin măng xông mới theo định dạng sau:\n\n"
            "TênMX,Latitude,Longitude\n"
            "Sau đó nhập lần lượt các cặp đấu nối (Đầu vào:Đầu ra), mỗi cặp trên 1 dòng.\n"
            "Nhập 'done' khi hoàn tất.\n\n"
            "Ví dụ:\n"
            "MX5,10.12345,106.12345\n"
            "1:1\n"
            "2:2\n"
            "...\n"
            "24:24\n"
            "done"
        )

        context.user_data['adding_mx'] = True
        return ADD_MX_NAME
    except Exception as e:
        logger.error(f"Error in add_mx command: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý lệnh /addmx.")
        return ConversationHandler.END


async def handle_add_mx_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý thông tin cơ bản của măng xông mới"""
    try:
        data = update.message.text.split(',')
        if len(data) != 3:
            await update.message.reply_text("Định dạng không đúng. Vui lòng nhập lại theo định dạng: TênMX,Lat,Long")
            return ADD_MX_NAME

        mx_name, lat, long = data
        mx_name = mx_name.strip().upper()

        try:
            lat = float(lat.strip())
            long = float(long.strip())
        except ValueError:
            await update.message.reply_text("Latitude và Longitude phải là số. Vui lòng nhập lại.")
            return ADD_MX_NAME

        if mx_name in CONNECTIONS:
            await update.message.reply_text(f"Măng xông {mx_name} đã tồn tại. Vui lòng chọn tên khác.")
            return ADD_MX_NAME

        # Lưu thông tin cơ bản
        context.user_data['new_mx'] = {
            'name': mx_name,
            'lat': lat,
            'long': long,
            'connections': {}
        }

        await update.message.reply_text(
            f"Đã nhận thông tin măng xông {mx_name} tại vị trí ({lat}, {long}).\n"
            "Vui lòng nhập các cặp đấu nối (Đầu vào:Đầu ra), mỗi cặp trên 1 dòng.\n"
            "Nhập 'done' khi hoàn tất."
        )
        return ADD_MX_CONNECTIONS
    except Exception as e:
        logger.error(f"Error in handle_add_mx_name: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý thông tin măng xông mới.")
        return ConversationHandler.END


async def handle_add_mx_connections(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý các cặp đấu nối của măng xông mới"""
    try:
        text = update.message.text.strip().lower()

        if text == 'done':
            # Kiểm tra đã có đủ 24 sợi chưa
            connections = context.user_data['new_mx']['connections']
            if len(connections) != 24:
                # Tạo danh sách các sợi đã nhập và chưa nhập
                entered_fibers = sorted(connections.keys())
                missing_fibers = [f for f in range(1, 25) if f not in connections]

                # Tạo thông báo chi tiết
                message = (
                    f"Bạn mới nhập được {len(connections)}/24 sợi.\n\n"
                    f"✅ Các sợi đã nhập: {', '.join(map(str, entered_fibers))}\n\n"
                    f"❌ Các sợi còn thiếu: {', '.join(map(str, missing_fibers))}\n\n"
                    "Vui lòng nhập tiếp các sợi còn thiếu hoặc nhập 'done' nếu muốn hủy."
                )

                await update.message.reply_text(message)
                return ADD_MX_CONNECTIONS

            # Thêm măng xông mới vào hệ thống
            mx_name = context.user_data['new_mx']['name']
            lat = context.user_data['new_mx']['lat']
            long = context.user_data['new_mx']['long']
            connections = context.user_data['new_mx']['connections']

            success = add_new_mx(mx_name, lat, long, connections)

            if success:
                await update.message.reply_text(
                    f"Đã thêm thành công măng xông {mx_name} vào hệ thống.\n"
                    f"Vị trí: {lat}, {long}\n"
                    "Bạn có thể tải file Excel mới nhất bằng lệnh /download."
                )
            else:
                await update.message.reply_text("Có lỗi xảy ra khi thêm măng xông mới.")

            # Xóa dữ liệu tạm
            if 'new_mx' in context.user_data:
                del context.user_data['new_mx']
            if 'adding_mx' in context.user_data:
                del context.user_data['adding_mx']

            return ConversationHandler.END

        # Xử lý cặp đấu nối
        if ':' not in text:
            await update.message.reply_text(
                "Định dạng không đúng. Vui lòng nhập theo định dạng: ĐầuVào:ĐầuRa (ví dụ: 1:1)")
            return ADD_MX_CONNECTIONS

        try:
            input_fiber, output_fiber = map(int, text.split(':'))
            if not (1 <= input_fiber <= 24 and 1 <= output_fiber <= 24):
                raise ValueError
        except ValueError:
            await update.message.reply_text("Số sợi phải từ 1 đến 24. Vui lòng nhập lại.")
            return ADD_MX_CONNECTIONS

        # Kiểm tra trùng lặp
        connections = context.user_data['new_mx']['connections']

        # Kiểm tra trùng sợi đầu vào
        if input_fiber in connections:
            await update.message.reply_text(
                f"Sợi đầu vào {input_fiber} đã được nhập trước đó. Vui lòng nhập lại."
            )
            return ADD_MX_CONNECTIONS

        # Kiểm tra trùng sợi đầu ra (trừ trường hợp đấu thẳng)
        if output_fiber in connections.values() and input_fiber != output_fiber:
            # Tìm sợi đầu vào nào đang sử dụng đầu ra này
            conflicting_input = next(
                (in_fib for in_fib, out_fib in connections.items() if out_fib == output_fiber),
                None
            )
            await update.message.reply_text(
                f"Sợi đầu ra {output_fiber} đã được sử dụng bởi sợi đầu vào {conflicting_input}. "
                "Vui lòng nhập lại."
            )
            return ADD_MX_CONNECTIONS

        # Lưu cặp đấu nối
        context.user_data['new_mx']['connections'][input_fiber] = output_fiber
        remaining = 24 - len(context.user_data['new_mx']['connections'])

        # Tạo danh sách cập nhật
        entered_fibers = sorted(context.user_data['new_mx']['connections'].keys())
        missing_fibers = [f for f in range(1, 25) if f not in context.user_data['new_mx']['connections']]

        # Tạo thông báo chi tiết
        message = (
            f"✅ Đã nhận cặp đấu nối {input_fiber}:{output_fiber}\n"
            f"📊 Tiến độ: {len(entered_fibers)}/24 sợi đã nhập\n\n"
            f"📌 Các sợi đã nhập: {', '.join(map(str, entered_fibers))}\n\n"
            f"🔍 Các sợi còn thiếu: {', '.join(map(str, missing_fibers))}\n\n"
            "Vui lòng nhập tiếp hoặc gõ 'done' để kết thúc"
        )

        await update.message.reply_text(message)
        return ADD_MX_CONNECTIONS
    except Exception as e:
        logger.error(f"Error in handle_add_mx_connections: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý thông tin đấu nối.")
        return ConversationHandler.END


# Thêm hàm cập nhật đấu nối trong dictionary CONNECTIONS
def update_mx_connections(mx_name, connections):
    """Cập nhật thông tin đấu nối của măng xông"""
    try:
        mx_name = mx_name.upper()
        if mx_name not in CONNECTIONS:
            return False

        CONNECTIONS[mx_name]['connections'] = connections
        return True
    except Exception as e:
        logger.error(f"Error in update_mx_connections: {e}")
        return False


# Thêm hàm cập nhật file Excel khi đấu nối mới
def update_excel_connections(mx_name, connections):
    """Cập nhật file Excel với thông tin đấu nối mới"""
    try:
        if not os.path.exists(MAIN_EXCEL_FILE):
            create_excel_file(MAIN_EXCEL_FILE)

        wb = openpyxl.load_workbook(MAIN_EXCEL_FILE)

        if mx_name not in wb.sheetnames:
            return False

        ws = wb[mx_name]

        # Cập nhật các cột đầu ra và ghi chú
        for fiber_num in range(1, 25):
            output_fiber = connections.get(fiber_num, fiber_num)
            note = 'Đấu thẳng' if fiber_num == output_fiber else 'Đấu chéo'

            # Cập nhật cột đầu ra (F) và ghi chú (G)
            ws.cell(row=fiber_num + 4, column=6).value = output_fiber  # Cột F
            ws.cell(row=fiber_num + 4, column=7).value = note  # Cột G

            # Định dạng lại màu cho ô đầu ra
            _, color_hex = FIBER_COLORS[output_fiber]
            fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
            ws.cell(row=fiber_num + 4, column=6).fill = fill
            text_color = '000000' if color_hex in ['FFFFFF', '00FFFF', 'FFFF00'] else 'FFFFFF'
            ws.cell(row=fiber_num + 4, column=6).font = Font(color=text_color)

        wb.save(MAIN_EXCEL_FILE)
        logger.info(f"Đã cập nhật file Excel với thông tin đấu nối mới cho {mx_name}")
        return True
    except Exception as e:
        logger.error(f"Error updating Excel connections for {mx_name}: {e}")
        return False


# Thêm hàm xử lý lệnh sửa măng xông
async def edit_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh sửa đấu nối măng xông"""
    try:
        user = update.effective_user

        # Kiểm tra quyền
        if not check_permission(user.username, 'write'):
            await update.message.reply_text(
                "Bạn không có quyền sửa măng xông. "
                "Liên hệ quản trị viên để được cấp quyền."
            )
            return ConversationHandler.END

        await update.message.reply_text(
            "Vui lòng nhập tên măng xông cần sửa đấu nối (ví dụ: MX1):"
        )
        return EDIT_MX
    except Exception as e:
        logger.error(f"Error in edit_mx command: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý lệnh sửa măng xông.")
        return ConversationHandler.END


# Thêm hàm xử lý tên măng xông cần sửa
async def handle_edit_mx(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý tên măng xông cần sửa"""
    try:
        mx_name = update.message.text.upper()

        if mx_name not in CONNECTIONS:
            await update.message.reply_text(
                f"Không tìm thấy măng xông {mx_name} trong hệ thống."
            )
            return ConversationHandler.END

        # Lưu tên măng xông vào context
        context.user_data['editing_mx'] = mx_name
        context.user_data['original_connections'] = CONNECTIONS[mx_name]['connections'].copy()

        # Hiển thị thông tin hiện tại và hướng dẫn
        message = (
            f"Thông tin đấu nối hiện tại của măng xông {mx_name}:\n\n"
            "Sợi | Đầu vào -> Đầu ra | Ghi chú\n"
            "---------------------------\n"
        )

        connections = CONNECTIONS[mx_name]['connections']
        for input_fiber, output_fiber in connections.items():
            note = "Thẳng" if input_fiber == output_fiber else "Chéo"
            color_name, _ = FIBER_COLORS[input_fiber]
            message += f"{input_fiber:2} ({color_name:10}) -> {output_fiber:2} | {note}\n"

        message += (
            "\nVui lòng nhập cặp đấu nối cần sửa theo định dạng:\n"
            "ĐầuVào:ĐầuRa (ví dụ: 1:2 để đổi đầu ra của sợi 1 thành 2)\n"
            "Nhập 'done' để kết thúc hoặc 'cancel' để hủy"
        )

        await update.message.reply_text(message)
        return EDIT_MX_CONNECTION
    except Exception as e:
        logger.error(f"Error in handle_edit_mx: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý yêu cầu sửa măng xông.")
        return ConversationHandler.END


# Thêm hàm xử lý sửa đấu nối
async def handle_edit_mx_connection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý thay đổi đấu nối của măng xông"""
    try:
        text = update.message.text.strip().lower()
        mx_name = context.user_data['editing_mx']
        connections = CONNECTIONS[mx_name]['connections'].copy()

        if text == 'done':
            # Cập nhật đấu nối mới vào hệ thống
            success = update_mx_connections(mx_name, connections)

            if success:
                # Cập nhật file Excel
                update_excel_connections(mx_name, connections)

                await update.message.reply_text(
                    f"Đã cập nhật thành công đấu nối cho măng xông {mx_name}.\n"
                    "Bạn có thể tải file Excel mới nhất bằng lệnh /download."
                )
            else:
                await update.message.reply_text("Có lỗi xảy ra khi cập nhật đấu nối.")

            # Xóa dữ liệu tạm
            if 'editing_mx' in context.user_data:
                del context.user_data['editing_mx']
            if 'original_connections' in context.user_data:
                del context.user_data['original_connections']

            return ConversationHandler.END

        if text == 'cancel':
            # Khôi phục lại đấu nối ban đầu nếu có
            if 'original_connections' in context.user_data:
                CONNECTIONS[mx_name]['connections'] = context.user_data['original_connections']

            # Xóa dữ liệu tạm
            if 'editing_mx' in context.user_data:
                del context.user_data['editing_mx']
            if 'original_connections' in context.user_data:
                del context.user_data['original_connections']

            await update.message.reply_text("Đã hủy thao tác sửa đấu nối.")
            return ConversationHandler.END

        # Xử lý cặp đấu nối
        if ':' not in text:
            await update.message.reply_text(
                "Định dạng không đúng. Vui lòng nhập theo định dạng: ĐầuVào:ĐầuRa (ví dụ: 1:2)")
            return EDIT_MX_CONNECTION

        try:
            input_fiber, output_fiber = map(int, text.split(':'))
            if not (1 <= input_fiber <= 24 and 1 <= output_fiber <= 24):
                raise ValueError
        except ValueError:
            await update.message.reply_text("Số sợi phải từ 1 đến 24. Vui lòng nhập lại.")
            return EDIT_MX_CONNECTION

        # Kiểm tra xem sợi đầu vào có tồn tại không
        if input_fiber not in connections:
            await update.message.reply_text(
                f"Sợi đầu vào {input_fiber} không tồn tại trong măng xông {mx_name}. Vui lòng nhập lại.")
            return EDIT_MX_CONNECTION

        # Kiểm tra xem sợi đầu ra đã được sử dụng bởi sợi khác chưa (trừ trường hợp đấu thẳng)
        if output_fiber in connections.values() and input_fiber != output_fiber:
            # Tìm sợi đầu vào nào đang sử dụng đầu ra này
            conflicting_input = next(
                (in_fib for in_fib, out_fib in connections.items() if out_fib == output_fiber),
                None
            )
            await update.message.reply_text(
                f"Sợi đầu ra {output_fiber} đã được sử dụng bởi sợi đầu vào {conflicting_input}. "
                "Vui lòng nhập lại."
            )
            return EDIT_MX_CONNECTION

        # Cập nhật đấu nối
        connections[input_fiber] = output_fiber
        CONNECTIONS[mx_name]['connections'] = connections  # Cập nhật tạm thời

        # Hiển thị thông tin cập nhật
        note = "Thẳng" if input_fiber == output_fiber else "Chéo"
        color_name, _ = FIBER_COLORS[input_fiber]

        await update.message.reply_text(
            f"Đã cập nhật: {input_fiber} ({color_name}) -> {output_fiber} | {note}\n\n"
            "Tiếp tục nhập cặp đấu nối khác cần sửa hoặc nhập 'done' để kết thúc."
        )
        return EDIT_MX_CONNECTION
    except Exception as e:
        logger.error(f"Error in handle_edit_mx_connection: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi xử lý yêu cầu sửa đấu nối.")
        return ConversationHandler.END

async def download(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lệnh tải file Excel"""
    try:
        filename = create_excel_file(MAIN_EXCEL_FILE)  # Đã chuyển thành synchronous function
        abs_path = os.path.abspath(filename)

        logger.info(f"Đang chuẩn bị tải file từ: {abs_path}")
        print(f"Đang chuẩn bị tải file từ: {abs_path}")

        with open(filename, 'rb') as file:
            await update.message.reply_document(
                document=file,
                caption=f"File Excel tổng hợp thông tin măng xông cáp quang\nĐường dẫn: {abs_path}"
            )

        logger.info(f"Đã gửi file thành công từ: {abs_path}")
        print(f"Đã gửi file thành công từ: {abs_path}")
    except Exception as e:
        error_msg = f"Error generating Excel file: {e}"
        logger.error(error_msg)
        print(error_msg)
        await update.message.reply_text("Có lỗi xảy ra khi tạo file Excel. Vui lòng thử lại sau.")


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Hủy bỏ conversation hiện tại"""
    try:
        # Xóa dữ liệu tạm nếu có
        if 'new_mx' in context.user_data:
            del context.user_data['new_mx']
        if 'adding_mx' in context.user_data:
            del context.user_data['adding_mx']

        await update.message.reply_text('Đã hủy thao tác hiện tại.')
    except Exception as e:
        logger.error(f"Error in cancel: {e}")
        if update.message:
            await update.message.reply_text("Có lỗi xảy ra khi hủy thao tác.")

    return ConversationHandler.END

async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Xử lý lỗi"""
    logger.error(f"Update {update} caused error {context.error}", exc_info=True)

    try:
        if update and update.message:
            await update.message.reply_text(
                "Có lỗi nghiêm trọng xảy ra trong hệ thống. "
                "Vui lòng thử lại hoặc liên hệ quản trị viên."
            )
    except Exception as e:
        logger.error(f"Error in error_handler: {e}")


def main():
    """Khởi chạy bot"""
    try:
        # Tạo file Excel ban đầu nếu chưa có
        if not os.path.exists(MAIN_EXCEL_FILE):
            create_excel_file(MAIN_EXCEL_FILE)

        # Tạo file phân quyền mẫu nếu chưa có
        if not os.path.exists(PERMISSION_FILE):
            permission_path = os.path.abspath(PERMISSION_FILE)
            logger.info(f"Đang tạo file phân quyền tại: {permission_path}")
            print(f"Đang tạo file phân quyền tại: {permission_path}")

            df = pd.DataFrame({'username': ['admin'], 'permission': ['write']})
            df.to_excel(PERMISSION_FILE, index=False)

            logger.info(f"Đã tạo file phân quyền thành công tại: {permission_path}")
            print(f"Đã tạo file phân quyền thành công tại: {permission_path}")

        # Lấy token từ biến môi trường hoặc nhập trực tiếp
        TOKEN = os.getenv('TELEGRAM_BOT_TOKEN') or '6183270075:AAHgGhIT5mjREJjyneaY9oLyxYJjhsJn36A'

        # Tạo application
        application = Application.builder().token(TOKEN).build()

        # Tạo ConversationHandler cho các lệnh
        conv_handler = ConversationHandler(
            entry_points=[
                CommandHandler('findmx', find_mx),
                CommandHandler('getmx', get_mx),
                CommandHandler('addmx', add_mx),
                CommandHandler('editmx', edit_mx)  # Thêm entry point mới
            ],
            states={
                FIND_MX: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_find_mx)],
                GET_MX: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_get_mx)],
                ADD_MX_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_add_mx_name)],
                ADD_MX_CONNECTIONS: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_add_mx_connections)],
                EDIT_MX_CONNECTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_edit_mx_connection)] # Thêm state mới
            },
            fallbacks=[CommandHandler('cancel', cancel)]
        )

        # Đăng ký các handler
        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("help", help_command))
        application.add_handler(CommandHandler("download", download))
        application.add_handler(conv_handler)

        # Đăng ký error handler
        application.add_error_handler(error_handler)

        # Khởi động bot
        application.run_polling()

    except Exception as e:
        logger.critical(f"Fatal error in main: {e}", exc_info=True)
        raise


if __name__ == '__main__':
    main()
